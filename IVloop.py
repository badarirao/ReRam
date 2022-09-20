# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'IVloop.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from winsound import MessageBeep
from itertools import chain
from PyQt5.QtCore import Qt, QObject, QThread, pyqtSignal
from time import sleep
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtWidgets import QMessageBox
from pyqtgraph import PlotWidget, ViewBox, mkPen, intColor
from numpy import linspace, array_split, around, concatenate, append
from numpy import reshape, array, savetxt, hsplit, insert, loadtxt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
from utilities import unique_filename, FakeAdapter, checkInstrument
from utilities import connect_sample_with_SMU, datetime

class Ui_IVLoop(QtWidgets.QWidget):
    """The pyqt5 gui class for IV loop measurement."""

    def __init__(self, parent=None):
        super(Ui_IVLoop, self).__init__(parent, QtCore.Qt.Window)
        self.setupUi(self)
        
    def setupUi(self, IVLoop):
        IVLoop.setObjectName("IVLoop")
        IVLoop.resize(818, 567)
        IVLoop.setMinimumSize(QtCore.QSize(0, 0))
        self.gridLayout_2 = QtWidgets.QGridLayout(IVLoop)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.groupBox = QtWidgets.QGroupBox(IVLoop)
        self.groupBox.setMinimumSize(QtCore.QSize(240, 260))
        self.groupBox.setMaximumSize(QtCore.QSize(242, 600))
        self.groupBox.setObjectName("groupBox")
        self.gridLayout = QtWidgets.QGridLayout(self.groupBox)
        self.gridLayout.setObjectName("gridLayout")
        self.temperature = QtWidgets.QDoubleSpinBox(self.groupBox)
        self.temperature.setEnabled(False)
        self.temperature.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.temperature.setMaximum(600.0)
        self.temperature.setProperty("value", 300.0)
        self.temperature.setObjectName("temperature")
        self.gridLayout.addWidget(self.temperature, 12, 2, 1, 1)
        self.ncycles_label = QtWidgets.QLabel(self.groupBox)
        self.ncycles_label.setObjectName("ncycles_label")
        self.gridLayout.addWidget(self.ncycles_label, 9, 0, 1, 2)
        self.Ilimit_label = QtWidgets.QLabel(self.groupBox)
        self.Ilimit_label.setObjectName("Ilimit_label")
        self.gridLayout.addWidget(self.Ilimit_label, 7, 0, 1, 2)
        self.scan_speed_label = QtWidgets.QLabel(self.groupBox)
        self.scan_speed_label.setObjectName("scan_speed_label")
        self.gridLayout.addWidget(self.scan_speed_label, 6, 0, 1, 2)
        self.temp_check = QtWidgets.QCheckBox(self.groupBox)
        self.temp_check.setObjectName("temp_check")
        self.gridLayout.addWidget(self.temp_check, 12, 0, 1, 2)
        self.setting_label = QtWidgets.QLabel(self.groupBox)
        self.setting_label.setMinimumSize(QtCore.QSize(0, 24))
        self.setting_label.setMaximumSize(QtCore.QSize(16777215, 50))
        self.setting_label.setObjectName("setting_label")
        self.gridLayout.addWidget(self.setting_label, 0, 0, 1, 3)
        self.Ilimit = QtWidgets.QDoubleSpinBox(self.groupBox)
        self.Ilimit.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.Ilimit.setDecimals(3)
        self.Ilimit.setMaximum(1.0)
        self.Ilimit.setSingleStep(0.001)
        self.Ilimit.setProperty("value", 0.5)
        self.Ilimit.setObjectName("Ilimit")
        self.gridLayout.addWidget(self.Ilimit, 7, 2, 1, 1)
        self.file_name = QtWidgets.QLineEdit(self.groupBox)
        self.file_name.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.file_name.setObjectName("file_name")
        self.gridLayout.addWidget(self.file_name, 1, 2, 1, 1)
        self.fname_label = QtWidgets.QLabel(self.groupBox)
        self.fname_label.setObjectName("fname_label")
        self.gridLayout.addWidget(self.fname_label, 1, 0, 1, 2)
        self.ncycles = QtWidgets.QSpinBox(self.groupBox)
        self.ncycles.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.ncycles.setMinimum(1)
        self.ncycles.setMaximum(500)
        self.ncycles.setProperty("value", 1)
        self.ncycles.setObjectName("ncycles")
        self.gridLayout.addWidget(self.ncycles, 9, 2, 1, 1)
        self.delay = QtWidgets.QDoubleSpinBox(self.groupBox)
        self.delay.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.delay.setMaximum(10000.0)
        self.delay.setSingleStep(0.05)
        self.delay.setProperty("value", 1.0)
        self.delay.setObjectName("delay")
        self.gridLayout.addWidget(self.delay, 5, 2, 1, 1)
        self.delay_label = QtWidgets.QLabel(self.groupBox)
        self.delay_label.setObjectName("delay_label")
        self.gridLayout.addWidget(self.delay_label, 5, 0, 1, 1)
        self.scan_speed = QtWidgets.QComboBox(self.groupBox)
        self.scan_speed.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.scan_speed.setObjectName("scan_speed")
        self.scan_speed.addItem("")
        self.scan_speed.addItem("")
        self.scan_speed.addItem("")
        self.scan_speed.addItem("")
        self.scan_speed.addItem("")
        self.gridLayout.addWidget(self.scan_speed, 6, 2, 1, 1)
        self.maxV_label = QtWidgets.QLabel(self.groupBox)
        self.maxV_label.setObjectName("maxV_label")
        self.gridLayout.addWidget(self.maxV_label, 4, 0, 1, 2)
        self.maxV = QtWidgets.QDoubleSpinBox(self.groupBox)
        self.maxV.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.maxV.setDecimals(3)
        self.maxV.setMinimum(-9.0)
        self.maxV.setMaximum(10.0)
        self.maxV.setSingleStep(0.001)
        self.maxV.setProperty("value", 3.0)
        self.maxV.setObjectName("maxV")
        self.gridLayout.addWidget(self.maxV, 4, 2, 1, 1)
        self.minV_label = QtWidgets.QLabel(self.groupBox)
        self.minV_label.setObjectName("minV_label")
        self.gridLayout.addWidget(self.minV_label, 3, 0, 1, 2)
        self.minV = QtWidgets.QDoubleSpinBox(self.groupBox)
        self.minV.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.minV.setDecimals(3)
        self.minV.setMinimum(-10.0)
        self.minV.setMaximum(9.0)
        self.minV.setSingleStep(0.001)
        self.minV.setProperty("value", -3.0)
        self.minV.setObjectName("minV")
        self.gridLayout.addWidget(self.minV, 3, 2, 1, 1)
        self.ncycles_label_2 = QtWidgets.QLabel(self.groupBox)
        self.ncycles_label_2.setObjectName("ncycles_label_2")
        self.gridLayout.addWidget(self.ncycles_label_2, 8, 0, 1, 1)
        self.ncycles_2 = QtWidgets.QSpinBox(self.groupBox)
        self.ncycles_2.setStyleSheet("font: 12pt \"Times New Roman\";")
        self.ncycles_2.setMinimum(2)
        self.ncycles_2.setMaximum(1000)
        self.ncycles_2.setProperty("value", 50)
        self.ncycles_2.setObjectName("ncycles_2")
        self.gridLayout.addWidget(self.ncycles_2, 8, 2, 1, 1)
        self.gridLayout_2.addWidget(self.groupBox, 0, 0, 1, 1)
        self.widget = QtWidgets.QWidget(IVLoop)
        self.widget.setMinimumSize(QtCore.QSize(240, 200))
        self.widget.setMaximumSize(QtCore.QSize(242, 10000))
        self.widget.setObjectName("widget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.widget)
        self.verticalLayout.setObjectName("verticalLayout")
        self.comment_checkBox = QtWidgets.QCheckBox(self.widget)
        self.comment_checkBox.setObjectName("comment_checkBox")
        self.verticalLayout.addWidget(self.comment_checkBox)
        self.commentBox = QtWidgets.QTextEdit(self.widget)
        self.commentBox.setEnabled(False)
        self.commentBox.setMaximumSize(QtCore.QSize(242, 16777215))
        self.commentBox.setObjectName("commentBox")
        self.verticalLayout.addWidget(self.commentBox)
        self.start_Button = QtWidgets.QPushButton(self.widget)
        self.start_Button.setStyleSheet("font: 75 14pt \"Times New Roman\";")
        self.start_Button.setObjectName("start_Button")
        self.verticalLayout.addWidget(self.start_Button)
        self.stop_Button = QtWidgets.QPushButton(self.widget)
        self.stop_Button.setStyleSheet("font: 75 14pt \"Times New Roman\";")
        self.stop_Button.setObjectName("stop_Button")
        self.verticalLayout.addWidget(self.stop_Button)
        self.statusbar = QtWidgets.QLineEdit(self.widget)
        self.statusbar.setEnabled(False)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(170, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        self.statusbar.setPalette(palette)
        self.statusbar.setStyleSheet("font: 11pt \"Times New Roman\";")
        self.statusbar.setText("")
        self.statusbar.setReadOnly(True)
        self.statusbar.setObjectName("statusbar")
        self.verticalLayout.addWidget(self.statusbar)
        self.gridLayout_2.addWidget(self.widget, 2, 0, 1, 1)
        self.graphWidget = PlotWidget(IVLoop, viewBox=ViewBox(border=mkPen(color='k', width=2)))
        self.graphWidget.setBackground((255, 182, 193, 25))
        self.graphWidget.setMinimumSize(QtCore.QSize(411, 379))
        self.graphWidget.setObjectName("graphWidget")
        self.gridLayout_2.addWidget(self.graphWidget, 0, 1, 3, 1)

        self.retranslateUi(IVLoop)
        self.scan_speed.setCurrentIndex(3)
        QtCore.QMetaObject.connectSlotsByName(IVLoop)
        IVLoop.setTabOrder(self.file_name, self.minV)
        IVLoop.setTabOrder(self.minV, self.maxV)
        IVLoop.setTabOrder(self.maxV, self.delay)
        IVLoop.setTabOrder(self.delay, self.scan_speed)
        IVLoop.setTabOrder(self.scan_speed, self.Ilimit)
        IVLoop.setTabOrder(self.Ilimit, self.ncycles_2)
        IVLoop.setTabOrder(self.ncycles_2, self.ncycles)
        IVLoop.setTabOrder(self.ncycles, self.temp_check)
        IVLoop.setTabOrder(self.temp_check, self.temperature)
        IVLoop.setTabOrder(self.temperature, self.comment_checkBox)
        IVLoop.setTabOrder(self.comment_checkBox, self.commentBox)
        IVLoop.setTabOrder(self.commentBox, self.start_Button)
        IVLoop.setTabOrder(self.start_Button, self.stop_Button)
        IVLoop.setTabOrder(self.stop_Button, self.statusbar)

    def retranslateUi(self, IVLoop):
        _translate = QtCore.QCoreApplication.translate
        IVLoop.setWindowTitle(_translate("IVLoop", "IV Loop"))
        self.temperature.setToolTip(_translate("IVLoop", "<html><head/><body><p>Temperature range will depend on the type of heater</p></body></html>"))
        self.ncycles_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Number of Cycles</span></p></body></html>"))
        self.Ilimit_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Current Limit (mA)</span></p></body></html>"))
        self.scan_speed_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Scan Speed</span></p></body></html>"))
        self.temp_check.setToolTip(_translate("IVLoop", "<html><head/><body><p>Use temperature only if temperature controller is attached</p></body></html>"))
        self.temp_check.setText(_translate("IVLoop", "Temperature (K)"))
        self.setting_label.setText(_translate("IVLoop", "<html><head/><body><p align=\"center\"><span style=\" font-size:12pt; font-weight:600;\">Parameters</span></p></body></html>"))
        self.Ilimit.setToolTip(_translate("IVLoop", "<html><head/><body><p>The compliance current, which protects the sample from full breakdown</p></body></html>"))
        self.file_name.setText(_translate("IVLoop", "Sample_IV"))
        self.fname_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">File Name</span></p></body></html>"))
        self.ncycles.setToolTip(_translate("IVLoop", "<html><head/><body><p>Number of cycles to be executed</p></body></html>"))
        self.delay_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Delay per point (ms)</span></p></body></html>"))
        self.scan_speed.setItemText(0, _translate("IVLoop", "Very Slow"))
        self.scan_speed.setItemText(1, _translate("IVLoop", "Slow"))
        self.scan_speed.setItemText(2, _translate("IVLoop", "Normal"))
        self.scan_speed.setItemText(3, _translate("IVLoop", "Fast"))
        self.scan_speed.setItemText(4, _translate("IVLoop", "Very Fast"))
        self.maxV_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Maximum Voltage (V)</span></p></body></html>"))
        self.maxV.setToolTip(_translate("IVLoop", "<html><head/><body><p>Max 10 V</p></body></html>"))
        self.minV_label.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Minimum Voltage (V)</span></p></body></html>"))
        self.minV.setToolTip(_translate("IVLoop", "<html><head/><body><p>Max -10 V</p></body></html>"))
        self.ncycles_label_2.setText(_translate("IVLoop", "<html><head/><body><p><span style=\" font-size:10pt;\">Number of Points</span></p></body></html>"))
        self.ncycles_2.setToolTip(_translate("IVLoop", "<html><head/><body><p>Number of cycles to be executed</p></body></html>"))
        self.comment_checkBox.setText(_translate("IVLoop", "Add Comments"))
        self.start_Button.setToolTip(_translate("IVLoop", "<html><head/><body><p>Click to start the experiment</p></body></html>"))
        self.start_Button.setText(_translate("IVLoop", "START"))
        self.stop_Button.setToolTip(_translate("IVLoop", "<html><head/><body><p>Click to abort the experiment</p></body></html>"))
        self.stop_Button.setText(_translate("IVLoop", "STOP"))

class app_IVLoop(Ui_IVLoop):
    """The IV-Loop app module."""

    def __init__(self, parent=None, smu=None, k2700 = None, sName="Sample_IV.txt",connection=1,currentSample=0):
        super(app_IVLoop, self).__init__(parent)
        self.parent = parent
        self.file_name.setReadOnly(True)
        self.smu = smu
        self.k2700 = k2700
        self.connection = connection
        self.currentSample = currentSample
        self.smu.reset()
        self.stop_Button.setEnabled(False)
        self.stop_flag = False
        self.start_Button.clicked.connect(self.start_ivloop)
        self.stop_Button.clicked.connect(self.stop_ivloop)
        self.start_Button.setShortcut('Ctrl+Return')
        self.stop_Button.setShortcut('Ctrl+q')
        self.initialize_plot()
        self.filename = sName
        self.file_name.setText(self.filename)
        self.measurement_status = "Idle"
        self.params = {
            "Vmin": -3,
            "Vmax": 3,
            "Delay": 50/1000,
            "Speed": 3,
            "ILimit": 1/1000,
            "ncycles": 1,
            "temperature": 300,
            "temp_check": 0,
            "comments": ""}
        self.parameters = list(self.params.values())[:-1]
        self.comment_checkBox.stateChanged.connect(self.updateCommentBox)
    
    def updateCommentBox(self):
        if self.comment_checkBox.isChecked():
            self.commentBox.setEnabled(True)
        else:
            self.commentBox.setEnabled(False)

    def load_parameters(self):
        try:
            self.minV.setValue(self.parameters[0])
            self.maxV.setValue(self.parameters[1])
            self.delay.setValue(self.parameters[2]*1000)
            self.scan_speed.setCurrentIndex(self.parameters[3])
            self.Ilimit.setValue(self.parameters[4]*1000)
            self.ncycles.setValue(self.parameters[5])
            self.temperature.setValue(self.parameters[6])
            self.temp_check.setChecked(self.parameters[7])
        except Exception:
            pass
    
    def update_params(self):
        """
        Update the measurement parameters.

        Returns
        -------
        None.

        """
        maincomment = self.parent.commentBox.toPlainText()
        wholeComment = maincomment + '\n' + self.commentBox.toPlainText()
        formattedComment = ""
        for t in wholeComment.split('\n'):
            formattedComment += '##' + t + '\n'
        self.params = {
            "Vmin": self.minV.value(),
            "Vmax": self.maxV.value(),
            "Delay": self.delay.value()/1000,
            "Speed": self.scan_speed.currentIndex(),
            "ILimit": self.Ilimit.value()/1000,
            "ncycles": self.ncycles.value(),
            "temperature": self.temperature.value(),
            "temp_check": int(self.temp_check.isChecked()),
            "comments": formattedComment}
        self.parameters = list(self.params.values())[:-1]
        if self.params["Vmax"] < self.params["Vmin"]:
            t = self.params["Vmax"]
            self.params["Vmax"] = self.params["Vmin"]
            self.params["Vmin"] = t
            self.minV.setValue(self.params["Vmin"])
            self.maxV.setValue(self.params["Vmax"])

    def plot_IVloop(self, Data):
        """
        Plot the ith IV loop into the graph.

        Parameters
        ----------
        data : list
            I values for every applied Voltage
        i : int
            Loop number

        Returns
        -------
        None.

        """
        data = Data[0]
        i = Data[1]
        self.currentCycle = i
        volts, currents = hsplit(data, 2)
        volts = volts.flatten()
        currents = abs(currents.flatten())
        pen1 = mkPen(intColor(3*i, values=3), width=2)
        if self.currentCycle == self.previousCycle:
            self.data_line.setData(self.points[0:len(volts)], currents)
        else:
            self.data_line = self.graphWidget.plot(self.points[0:len(volts)], currents, name="Cycle {0}".format(i), pen=pen1)
        self.previousCycle = i

    def start_ivloop(self):
        """
        Start the IV measurement based on given parameters.

        Returns
        -------
        None.

        """
        self.previousCycle = 0
        self.fullfilename = unique_filename(directory='.', prefix=self.filename, ext='dat', datetimeformat="")
        self.statusbar.setText("Measurement Running..")
        self.measurement_status = "Running"
        if self.stop_flag:
            self.stop_flag = False
            self.graphWidget.clear()
        self.stop_Button.setEnabled(True)
        self.minV.setEnabled(False)
        self.maxV.setEnabled(False)
        self.delay.setEnabled(False)
        self.scan_speed.setEnabled(False)
        self.Ilimit.setEnabled(False)
        self.ncycles.setEnabled(False)
        self.temp_check.setEnabled(False)
        self.start_Button.setEnabled(False)
        self.update_params()
        self.startThread()

    def startThread(self):
        self.thread = QThread()
        self.worker = Worker(self.params, self.smu, self.k2700, self.fullfilename, self.connection, self.currentSample)
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.start_IV)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.worker.data.connect(self.plot_IVloop)
        self.thread.finished.connect(self.finishAction)
        self.worker.sendPoints.connect(self.getPoints)
        self.thread.finished.connect(self.finishAction)   
        self.thread.start()

    def getPoints(self,points):
        self.points = points[0]

    def finishAction(self):
        if not self.stop_flag:
            self.statusbar.setText("Measurement Finished.")
            self.measurement_status = "Idle"
            self.stop_flag = True
        self.minV.setEnabled(True)
        self.maxV.setEnabled(True)
        self.delay.setEnabled(True)
        self.scan_speed.setEnabled(True)
        self.Ilimit.setEnabled(True)
        self.ncycles.setEnabled(True)
        self.temp_check.setEnabled(True)
        self.start_Button.setEnabled(True)
        self.smu.source_voltage = 0
        self.stop_Button.setEnabled(False)
        app_IVLoop.formatIV_Excel(self.fullfilename)
    
    def initialize_plot(self):
        """
        Initialize the plot to display IV loop.

        Returns
        -------
        None.

        """
        styles = {'color': 'r', 'font-size': '20px'}
        self.graphWidget.setLabel('left', '|Current (A)|', **styles)
        self.graphWidget.setLabel('bottom', 'Voltage (V)', **styles)
        self.graphWidget.setRange(QtCore.QRectF(self.minV.value(
        ), 1e-12, self.maxV.value()-self.minV.value(), self.Ilimit.value()), padding=0)
        self.graphWidget.getPlotItem().setLogMode(None, True)
        self.graphWidget.addLegend()
        # self.graphWidget.plot((0,0),(1e-12,2),pen=pen1)
        # self.graphWidget.addLegend(offset=(180,170))
        #self.graphWidget.showGrid(x=True, y=True)

    def stop_ivloop(self):
        """
        Trigger to stop the IV measurement.

        Returns
        -------
        None.

        """
        self.statusbar.setText("Measurement Aborted!")
        self.measurement_status = "Aborted"
        self.stop_flag = True
        self.worker.stopcall.emit()

    # make sure filename has some extension, it should be the file name that has collected data
    @staticmethod
    def formatIV_Excel(fname="Sample_IV.dat"):
        """
        Save the data and the plot into excel file,.
        *** Obselete now *** may be removed later
        Parameters
        ----------
        fname : TYPE, optional
            DESCRIPTION. The default is "Sample_IV.txt".

        Returns
        -------
        None.

        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'IV Data'
        dataFile = loadtxt(fname)
        dataFile = append(dataFile, array([abs(dataFile[:, 2])]).T, axis=1)
        header1s = []
        header2s = []
        legend = []
        cycles = 0
        myfile = open(fname, "r")
        h = ''
        while myfile:
            line = myfile.readline()
            li = line.strip()
            if li.startswith('##'):
                continue
            if li.startswith('#'):
                line2 = myfile.readline()
                li2 = line2.strip()
                if li2.startswith('#'):
                    l = li2
                    li2 = li
                    li = l
                    li2 = li2[1:].split('\t')
                    li2.append('Absolute Current (A)')
                    h = li2
                header2s.append(h)
                legend.append(li)
                header1s.append([' ', ' ', ' ', li])
                line2 = myfile.readline()
                li2 = line2.strip()
                cycles = cycles + 1

            if line == "":
                break
        myfile.close()
        header1 = list(chain.from_iterable(header1s))
        header2 = list(chain.from_iterable(header2s))
        ws.append(header1)
        ws.append(header2)
        rownum = 3
        colnum = 1
        for r in ws.iter_rows(min_row=1, max_row=2, max_col=len(header2)):
            for cell in r:
                cell.font = Font(bold=True)
        c = 1
        for row in dataFile:
            for value in row:
                ws.cell(row=rownum, column=colnum).value = value
                colnum = colnum + 1
            colnum = c
            rownum = rownum+1
            if (rownum-3) % 501 == 0:
                rownum = 3
                c = c + 4
        ws.freeze_panes = 'A3'
        for column_cells in ws.columns:
            ws.column_dimensions[column_cells[0].column_letter].width = 20

        # plot all IV loops
        plotsheet = wb.create_sheet()
        plotsheet.title = 'IV Plot'
        ivChart = ScatterChart()
        ivChart.height = 14
        ivChart.width = 25
        ivChart.title = "Cumulative IV Plots (log-scale)"
        ivChart.style = 12
        ivChart.y_axis.title = "Abs. Current (A)"
        ivChart.y_axis.scaling.logBase = 10
        ivChart.y_axis.scaling.max = dataFile[:, 3].max()*1.2
        ivChart.x_axis.title = 'Voltage (V)'
        paraprops = ParagraphProperties()

        # change the x and y-axis title font size
        paraprops.defRPr = CharacterProperties(sz=1600)
        for para in ivChart.x_axis.title.tx.rich.paragraphs:
            para.pPr = paraprops
        for para in ivChart.y_axis.title.tx.rich.paragraphs:
            para.pPr = paraprops

        ivChart.x_axis.tickLblPos = "low"
        ivChart.y_axis.tickLblPos = "low"
        for i in range(int(ws.max_column/4)):
            xvalues = Reference(ws, min_col=i*4+2,
                                min_row=3, max_row=ws.max_row)
            yvalues = Reference(ws, min_col=i*4+4,
                                min_row=3, max_row=ws.max_row)
            series = Series(yvalues, xvalues,
                            title_from_data=False, title=legend[i])
            ivChart.series.append(series)
        plotsheet.add_chart(ivChart, 'C3')
        wb.save('.'.join(fname.split('.')[:-1])+'.xlsx')
        # wb.save(self.filename+'.xlsx')
    
    def keyPressEvent(self, event):
        """Close application from escape key.

        results in QMessageBox dialog from closeEvent, good but how/why?
        """
        if event.key() == Qt.Key_Escape:
            self.close()
    
    def closeEvent(self, event):
        """
        Perform necessary operations just before exiting the program.

        Parameters
        ----------
        event : QCloseEvent

        Returns
        -------
        None.

        """
        reply = QMessageBox.Yes
        if self.measurement_status == "Running":
            quit_msg = "Measurement is in Progress. Are you sure you want to stop and exit?"
            reply = QMessageBox.question(self, 'Message', quit_msg, QMessageBox.Yes, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.worker.stopcall.emit()
        if reply == QMessageBox.Yes:
            if __name__ != "__main__":
                self.parent.show()
            event.accept()
        else:
            event.ignore()

class Worker(QObject):
    finished = pyqtSignal()
    data = pyqtSignal(list)
    stopcall = pyqtSignal()
    sendPoints = pyqtSignal(list)
    
    def __init__(self, params, smu=None, k2700=None, fullfilename="sample.dat", connection = 1, currentSample = 0):
        super(Worker,self).__init__()
        self.stopCall = False
        self.params = params
        self.smu = smu
        self.k2700 = k2700
        self.connection = connection
        self.currentSample = currentSample
        self.fullfilename = fullfilename
        self.stopcall.connect(self.stopcalled)
        self.npoints = 100
        self.smu.nplc = 1
        self.status = 1
        
    def initialize_SMU(self):
        """
        Initialize the SMU.

        Returns
        -------
        None.

        """
        if self.smu is None:
            self.smu = FakeAdapter()
        if self.k2700 is None:
            self.k2700 = FakeAdapter()
        
        # if afg is connected, remove and connect the source meter
        connect_sample_with_SMU(self.k2700,self.connection,self.currentSample)
        self.smu.measure_current()
        self.smu.auto_range_sense()
        self.smu.set_wire_configuration(2) # two wire configuration
        if self.params["Speed"] == 0:
            nplc = 5
            self.speed = "Very Slow"
        elif self.params["Speed"] == 1:
            nplc = 2
            self.speed = "Slow"
        elif self.params["Speed"] == 2:
            nplc = 1
            self.speed = "Normal"
        elif self.params["Speed"] == 3:
            nplc = 0.1
            self.speed = "Fast"
        elif self.params["Speed"] == 4:
            nplc = 0.01
            self.speed = "Very Fast"
        self.smu.nplc = nplc
        self.smu.apply_voltage(compliance_current=self.params["ILimit"])
        # correct for zero only at the beginning
        self.smu.correct_zero_at_beginning_only()
        self.smu.set_read_back_on()

    def configure_sweep(self):
        """
        Configure the sweep conditions based on given parameters.

        Returns
        -------
        None.

        """
        with open(self.fullfilename,'w') as f:
            f.write("## IV loop measurement using Keithley 2450 source measure unit.\n")
            f.write(f"## Date & Time: {datetime.now().strftime('%m/%d/%Y, %H:%M:%S')}\n")
            f.write("## Min voltage = {0}V, Max voltage = {1}V\n".format(self.params["Vmin"],self.params["Vmax"]))
            f.write('## Limiting current = {0} mA, Delay per point = {1}ms\n'.format(self.params["ILimit"]*1000,self.params["Delay"]))
            f.write('## Scan speed = {0}, Requested number of IV loops = {1}\n'.format(self.speed,self.params["ncycles"]))
            f.write(self.params["comments"])
            f.write("#Set Voltage(V)\tActual Voltage(V)\tCurrent(A)\n")
            
        if self.params["Vmax"] == self.params["Vmin"]:
            self.points = [self.params["Vmax"]]
            self.smu.set_voltage_points(self.points[0])
        elif self.params["Vmax"] >= 0 >= self.params["Vmin"]:
            nplus = int(
                self.params["Vmax"]/(self.params["Vmax"]-self.params["Vmin"])*self.npoints*0.5)
            nminus = int(abs(
                self.params["Vmin"])/(self.params["Vmax"]-self.params["Vmin"])*self.npoints*0.5)
            l1 = linspace(0, self.params["Vmax"], nplus, endpoint=False)
            l2 = linspace(
                self.params["Vmax"], self.params["Vmin"], nplus+nminus, endpoint=False)
            l3 = linspace(self.params["Vmin"], 0, nminus+1, endpoint=True)
            self.points = around(concatenate((l1, l2, l3)), 3)
            self.points[self.points == 0] = 0.0001
            # split the points into 6 chunks of equal size
            self.chunks = array_split(self.points, 6)
            # write the first chunk into the list
            self.smu.set_voltage_points(str(list(self.chunks[0]))[1:-1])
            # append the remaining chunks into the list
            if len(self.chunks) > 1:
                for i in self.chunks[1:]:
                    self.smu.append_voltage_points(str(list(i))[1:-1])
        else:
            l1 = linspace(self.params["Vmin"], self.params["Vmax"], int(
                self.npoints/2), endpoint=False)
            l2 = linspace(self.params["Vmax"], self.params["Vmin"], int(
                self.npoints/2)+1, endpoint=True)
            self.points = around(concatenate((l1, l2)), 3)
            self.chunks = array_split(self.points, 5)
            # write the first chunk into the list
            self.smu.set_voltage_points(str(list(self.chunks[0]))[1:-1])
            # append the remaining chunks into the list
            if len(self.chunks) > 1:
                for i in self.chunks[1:]:
                    self.smu.append_voltage_points(str(list(i))[1:-1])
        # set the sweep function with the above list
        self.smu.configure_sweep(self.params["Delay"])
        self.sendPoints.emit([self.points])
    
    def start_IV(self):
        self.initialize_SMU()
        self.configure_sweep()
        i = 0
        while i<self.params["ncycles"] and not self.stopCall:
            self.smu.start_buffer()  # start the measurement
            while self.stopCall == False:
                sleep(1)
                start_point = self.smu.get_start_point()
                if start_point == 0:
                    continue
                end_point = self.smu.get_end_point()
                data = self.smu.get_trace_data(start_point, end_point)
                data = reshape(array(data.split(','), dtype=float), (-1, 2))
                self.data.emit([data,i+1])
                state = self.smu.get_trigger_state()
                if state != 'RUNNING':
                    if state != 'IDLE':
                        self.status = 0
                    break
            with open(self.fullfilename, "a") as f:
                f.write("#Cycle {0}\n".format(i+1))
                data = insert(data, 0, self.points[0:len(data)], axis=1)
                savetxt(f, data, delimiter='\t')
                f.write("\n\n")
            i = i + 1
        if self.stopCall:
            self.smu.abort()
        self.smu.source_voltage = 0
        self.smu.disable_source()
        MessageBeep()
        self.finished.emit()
    
    def stopcalled(self):
        self.stopCall = True

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    IVLoop = QtWidgets.QWidget()
    smu, k2700, _ = checkInstrument(test = True)
    ui = app_IVLoop(IVLoop,smu,k2700)
    ui.show()
    app.exec_()
    app.quit()